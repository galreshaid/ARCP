import csv
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import (
    MaterialCatalog,
    MaterialCategory,
    MaterialMeasurement,
    Procedure,
    ProcedureMaterialBundle,
    ProcedureMaterialBundleItem,
)


class Command(BaseCommand):
    help = "Import radiology consumables catalog and optional procedure material bundles from CSV/XLSX."

    def add_arguments(self, parser):
        parser.add_argument("materials_file", help="Path to materials catalog CSV/XLSX file.")
        parser.add_argument(
            "--bundles-file",
            default="",
            help="Optional path to procedure-bundle mapping CSV/XLSX file.",
        )
        parser.add_argument(
            "--materials-sheet",
            default="",
            help="Optional XLSX sheet name for materials import.",
        )
        parser.add_argument(
            "--bundles-sheet",
            default="",
            help="Optional XLSX sheet name for bundle import.",
        )
        parser.add_argument(
            "--deactivate-missing",
            action="store_true",
            help="Deactivate existing catalog rows not present in the imported material codes.",
        )

    def _normalize_header(self, value: str) -> str:
        return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum() or ch == "_")

    def _iter_csv_rows(self, path):
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as source:
            reader = csv.DictReader(source)
            for row in reader:
                if not row:
                    continue
                yield {self._normalize_header(k): str(v or "").strip() for k, v in row.items()}

    def _iter_xlsx_rows(self, path, sheet_name=""):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("openpyxl is required for XLSX import.") from exc

        workbook = load_workbook(path, data_only=True, read_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f"Sheet '{sheet_name}' not found in {path}.")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return

        headers = [self._normalize_header(value) for value in header_row]
        for values in rows:
            payload = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                payload[header] = str(values[index] if index < len(values) else "" or "").strip()
            if payload:
                yield payload

    def _iter_rows(self, path, sheet_name=""):
        suffix = Path(path).suffix.lower()
        if suffix == ".csv":
            yield from self._iter_csv_rows(path)
            return
        if suffix == ".xlsx":
            yield from self._iter_xlsx_rows(path, sheet_name=sheet_name)
            return
        raise CommandError("Unsupported file type. Use .csv or .xlsx.")

    def _row_value(self, row, *keys, default=""):
        for key in keys:
            normalized = self._normalize_header(key)
            if normalized in row:
                value = str(row.get(normalized) or "").strip()
                if value:
                    return value
        return default

    def _parse_decimal(self, value):
        raw = str(value or "").strip().replace(",", "")
        if not raw:
            return None
        try:
            return Decimal(raw)
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _parse_bool(self, value, *, default=False):
        raw = str(value or "").strip().lower()
        if not raw:
            return default
        if raw in {"1", "true", "yes", "y", "on"}:
            return True
        if raw in {"0", "false", "no", "n", "off"}:
            return False
        return default

    def _parse_int(self, value, *, default=0):
        try:
            return int(str(value or "").strip())
        except (TypeError, ValueError):
            return default

    def _first_number(self, value: str) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        match = re.search(r"(\d+(?:\.\d+)?)", text)
        return match.group(1) if match else ""

    def _optional_metadata(self, row):
        route = self._row_value(row, "Route", "route", "default_route")
        concentration_hint = self._row_value(
            row,
            "DefaultConcentrationMgMl",
            "default_concentration_mg_ml",
            "TypicalConcentration",
            "typical_concentration",
            "Concentration",
            "concentration",
            "FormStrength",
            "form_strength",
        )
        injection_rate_hint = self._row_value(
            row,
            "DefaultInjectionRateMlS",
            "default_injection_rate_ml_s",
            "InjectionRateMlS",
            "injection_rate_ml_s",
            "TypicalInjectionRateMlS",
            "typical_injection_rate_ml_s",
        )

        metadata = {
            "import_source": "import_consumables_catalog",
        }

        for metadata_key, keys in (
            ("generic_name", ("GenericName", "generic_name")),
            ("brand_name", ("BrandName", "brand_name")),
            ("form_strength", ("FormStrength", "form_strength")),
            ("route", ("Route", "route")),
            ("typical_adult_dose", ("TypicalAdultDose", "typical_adult_dose")),
            ("typical_peds_dose", ("TypicalPedsDose", "typical_peds_dose")),
            ("half_life", ("HalfLife", "half_life")),
            ("storage", ("Storage", "storage")),
            ("manufacturer", ("Manufacturer", "manufacturer")),
            ("contraindications", ("Contraindications", "contraindications", "contraindications_key")),
            ("indications", ("Indications", "indications")),
            ("osmolality", ("Osmolality", "osmolality")),
            ("department", ("Department", "department")),
            ("auto_consumption", ("AutoConsumption", "auto_consumption")),
            ("uom", ("UOM", "uom")),
        ):
            value = self._row_value(row, *keys)
            if value:
                metadata[metadata_key] = value

        default_concentration = self._first_number(concentration_hint)
        if default_concentration:
            metadata["default_concentration_mg_ml"] = default_concentration

        default_injection_rate = self._first_number(injection_rate_hint)
        if default_injection_rate:
            metadata["default_injection_rate_ml_s"] = default_injection_rate

        if route:
            metadata["default_route"] = str(route).strip().upper()

        return metadata

    def _measurement_for_unit(self, unit_code: str):
        code = str(unit_code or "").strip()
        if not code:
            return None
        measurement, _ = MaterialMeasurement.objects.get_or_create(
            code=code,
            defaults={
                "label": code,
                "is_active": True,
            },
        )
        if not measurement.is_active:
            measurement.is_active = True
            measurement.save(update_fields=["is_active"])
        return measurement

    def _normalized_category(self, raw_category: str) -> str:
        value = str(raw_category or "").strip()
        if not value:
            return MaterialCategory.DISPOSABLE
        return value

    def _import_materials(self, path, *, sheet_name="", deactivate_missing=False):
        if not Path(path).exists():
            raise CommandError(f"Materials file not found: {path}")

        created = 0
        updated = 0
        skipped = 0
        imported_codes = set()

        for row in self._iter_rows(path, sheet_name=sheet_name):
            material_code = self._row_value(row, "MaterialCode", "material_code")
            material_name = self._row_value(row, "MaterialName", "material_name", "name")
            if not material_code and not material_name:
                skipped += 1
                continue

            if material_code:
                lookup = {"material_code": material_code}
                imported_codes.add(material_code)
            else:
                lookup = {"name": material_name}

            unit = self._row_value(row, "Unit", "unit")
            measurement = self._measurement_for_unit(unit)

            defaults = {
                "name": material_name or material_code,
                "category": self._normalized_category(self._row_value(row, "Category", "category")),
                "unit": unit,
                "pack_size": self._row_value(row, "PackSize", "pack_size"),
                "modality_scope": self._row_value(row, "ModalityScope", "modality_scope"),
                "procedure_mapping_tags": self._row_value(
                    row,
                    "ProcedureMappingTags",
                    "procedure_mapping_tags",
                    "procedure_mapping",
                ),
                "charge_code": self._row_value(row, "ChargeCode", "charge_code"),
                "billing_ref_example": self._row_value(
                    row,
                    "BillingRef_Example",
                    "billingref_example",
                    "billing_ref_example",
                ),
                "nphies_code": self._row_value(row, "nphies_code", "NPHIESCode"),
                "typical_cost_sar": self._parse_decimal(
                    self._row_value(row, "TypicalCost_SAR", "typical_cost_sar", "cost_sar")
                ),
                "default_price_sar": self._parse_decimal(
                    self._row_value(row, "DefaultPrice_SAR", "default_price_sar", "charge_sar")
                ),
                "billable": self._parse_bool(
                    self._row_value(row, "billable", "Billable"),
                    default=True,
                ),
                "cost_center_only": self._parse_bool(
                    self._row_value(row, "cost_center_only", "CostCenterOnly"),
                    default=False,
                ),
                "reorder_level": self._parse_int(
                    self._row_value(row, "reorder_level", "ReorderLevel"),
                    default=0,
                ),
                "notes": self._row_value(row, "Notes", "notes"),
                "default_measurement": measurement,
                "is_active": True,
                "metadata": self._optional_metadata(row),
            }

            obj, is_created = MaterialCatalog.objects.update_or_create(
                defaults=defaults,
                **lookup,
            )

            if not obj.material_code and material_code:
                obj.material_code = material_code
                obj.save(update_fields=["material_code"])

            if is_created:
                created += 1
            else:
                updated += 1

        deactivated = 0
        if deactivate_missing and imported_codes:
            deactivated = (
                MaterialCatalog.objects.filter(is_active=True)
                .exclude(material_code__in=imported_codes)
                .exclude(material_code__isnull=True)
                .update(is_active=False)
            )

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "deactivated": deactivated,
        }

    def _parse_bundle_items(self, raw_value: str):
        items = []
        for sort_order, chunk in enumerate(str(raw_value or "").split(";"), start=1):
            token = str(chunk or "").strip()
            if not token:
                continue
            if ":" not in token:
                code = token
                quantity = Decimal("1.000")
            else:
                code_part, quantity_part = token.split(":", 1)
                code = str(code_part or "").strip()
                parsed_qty = self._parse_decimal(quantity_part)
                quantity = parsed_qty if parsed_qty is not None else Decimal("1.000")

            if not code:
                continue

            items.append(
                {
                    "material_code": code,
                    "quantity": quantity,
                    "sort_order": sort_order * 10,
                }
            )
        return items

    def _import_bundles(self, path, *, sheet_name=""):
        if not Path(path).exists():
            raise CommandError(f"Bundle file not found: {path}")

        created = 0
        updated = 0
        skipped = 0
        unresolved_items = 0

        for row in self._iter_rows(path, sheet_name=sheet_name):
            procedure_code = self._row_value(row, "ProcedureCode", "procedure_code")
            if not procedure_code:
                skipped += 1
                continue

            procedure_name = self._row_value(row, "ProcedureName", "procedure_name")
            modality = self._row_value(row, "Modality", "modality")
            bundle_items = self._row_value(
                row,
                "BundleItems (MaterialCode:Qty)",
                "BundleItems",
                "bundle_items",
            )
            rules_notes = self._row_value(row, "Rules/Notes", "rules_notes", "notes")

            procedure = Procedure.objects.filter(code=procedure_code).first()
            bundle, is_created = ProcedureMaterialBundle.objects.update_or_create(
                procedure_code=procedure_code,
                defaults={
                    "procedure": procedure,
                    "procedure_name": procedure_name,
                    "modality_scope": modality,
                    "rules_notes": rules_notes,
                    "is_active": True,
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1

            bundle.items.all().delete()
            parsed_items = self._parse_bundle_items(bundle_items)
            for item in parsed_items:
                material_code = item["material_code"]
                material = MaterialCatalog.objects.filter(material_code__iexact=material_code).first()
                if material is None:
                    unresolved_items += 1
                ProcedureMaterialBundleItem.objects.create(
                    bundle=bundle,
                    material=material,
                    material_code=material_code,
                    quantity=item["quantity"],
                    sort_order=item["sort_order"],
                )

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "unresolved_items": unresolved_items,
        }

    @transaction.atomic
    def handle(self, *args, **options):
        materials_file = options["materials_file"]
        bundles_file = str(options.get("bundles_file") or "").strip()
        materials_sheet = str(options.get("materials_sheet") or "").strip()
        bundles_sheet = str(options.get("bundles_sheet") or "").strip()
        deactivate_missing = bool(options.get("deactivate_missing"))

        material_summary = self._import_materials(
            materials_file,
            sheet_name=materials_sheet,
            deactivate_missing=deactivate_missing,
        )
        self.stdout.write(
            self.style.SUCCESS(
                "Materials import finished: "
                f"{material_summary['created']} created, "
                f"{material_summary['updated']} updated, "
                f"{material_summary['skipped']} skipped, "
                f"{material_summary['deactivated']} deactivated."
            )
        )

        if bundles_file:
            bundle_summary = self._import_bundles(bundles_file, sheet_name=bundles_sheet)
            self.stdout.write(
                self.style.SUCCESS(
                    "Bundle import finished: "
                    f"{bundle_summary['created']} created, "
                    f"{bundle_summary['updated']} updated, "
                    f"{bundle_summary['skipped']} skipped, "
                    f"{bundle_summary['unresolved_items']} unresolved bundle item codes."
                )
            )
